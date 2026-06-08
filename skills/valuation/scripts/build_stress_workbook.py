"""Build the BBCA stress-test Excel workbook.

Tabs:
  1. Drivers      - locked assumptions (yellow inputs) + the year-by-year driver
                    tree (NI, ROE, equity cash flow) + base outputs.
  2. Equity DCF   - LIVE formulas: discount the equity cash flows at ke, add the
                    value-driver continuing value, get value per share. Edit ke /
                    g / RONE / cash flows and it recalcs.
  3. Tornado      - drivers ranked by swing in value per share + a bar chart.
  4. Reverse DCF  - the driver value implied by today's price.
  5. Monte Carlo  - percentiles, probability undervalued, and a histogram chart.

Run with the skill venv (needs openpyxl):
    .venv/bin/python build_stress_workbook.py <drivers.json> --stress <stress.json>
        [-o out.xlsx] [--n 10000] [--seed 42]
"""

from __future__ import annotations

import argparse
import json

import driver_bank
import engine
import stress

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="D6DCE5")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
RESULT_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
ITAL = Font(italic=True, color="808080")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT = "0.00%"
NUM = "#,##0"
CENTER = Alignment(horizontal="center")


def _hdr(ws, row, cols):
    for j, txt in enumerate(cols, start=1):
        c = ws.cell(row=row, column=j, value=txt)
        c.fill = HEADER_FILL
        c.font = WHITE
        c.alignment = CENTER


def _section(ws, row, text, span=4):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL


# --------------------------------------------------------------------------
# Sheet 1 - Drivers
# --------------------------------------------------------------------------

def sheet_drivers(ws, drivers, fc, result):
    cur = drivers.get("currency", "IDR")
    ws.column_dimensions["A"].width = 30
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 13

    ws["A1"] = f"{drivers['company']} ({drivers['ticker']}) - Stress-test drivers"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"billions {cur} | as of {drivers.get('valuation_date','')}"
    ws["A2"].font = ITAL

    coe = drivers["cost_of_equity"]
    cv = drivers["continuing_value"]
    d = drivers["drivers"]
    ke = coe["risk_free"] + coe.get("beta", 1.0) * coe["erp"]

    r = 4
    _section(ws, r, "Locked assumptions", 2); r += 1
    assumptions = [
        ("Risk-free rate", coe["risk_free"], PCT),
        ("Equity risk premium", coe["erp"], PCT),
        ("Beta", coe.get("beta", 1.0), "0.00"),
        ("Cost of equity (ke)", ke, PCT),
        ("Terminal growth (g)", cv["growth"], PCT),
        ("RONE", cv["rone"], PCT),
        ("NIM", d["nim"], PCT),
        ("Cost of credit", d["cost_of_credit"], PCT),
        ("Cost-to-income", d["cost_to_income"], PCT),
        ("Tax rate", d["tax_rate"], PCT),
        ("Equity / assets", d["equity_to_assets"], PCT),
    ]
    for name, val, fmt in assumptions:
        ws[f"A{r}"] = name
        c = ws[f"B{r}"]
        c.value = val
        c.number_format = fmt
        c.fill = INPUT_FILL
        c.border = BOX
        r += 1

    r += 1
    _section(ws, r, "Driver tree (forecast)", 9); r += 1
    _hdr(ws, r, ["Year", "Growth", "Earn assets", "Net int inc",
                 "Op income", "Provisions", "Net income", "ROE", "Equity CF"]); r += 1
    for row in fc["rows"]:
        ws.cell(row=r, column=1, value=row["year"])
        ws.cell(row=r, column=2, value=row["growth"]).number_format = PCT
        ws.cell(row=r, column=3, value=round(row["earning_assets"])).number_format = NUM
        ws.cell(row=r, column=4, value=round(row["net_interest_income"])).number_format = NUM
        ws.cell(row=r, column=5, value=round(row["operating_income"])).number_format = NUM
        ws.cell(row=r, column=6, value=round(row["provisions"])).number_format = NUM
        ws.cell(row=r, column=7, value=round(row["net_income"])).number_format = NUM
        ws.cell(row=r, column=8, value=row["roe"]).number_format = PCT
        ws.cell(row=r, column=9, value=round(row["equity_cash_flow"])).number_format = NUM
        r += 1

    r += 1
    _section(ws, r, "Base outputs", 2); r += 1
    ps = result["value_per_share"]
    price = drivers.get("price")
    outs = [
        ("Implied steady-state ROE", fc["steady_state_roe"], PCT),
        ("NI next (CV year)", round(fc["ni_next"]), NUM),
        ("PV explicit equity CF", round(result["pv_explicit_fcf"]), NUM),
        ("PV continuing value", round(result["continuing_value"]["pv"]), NUM),
        ("Continuing value share", result["continuing_value"]["share_of_operations"], PCT),
        ("Equity value", round(result["equity_value"]), NUM),
        (f"Value per share ({cur})", round(ps), NUM),
    ]
    for name, val, fmt in outs:
        ws[f"A{r}"] = name
        c = ws[f"B{r}"]
        c.value = val
        c.number_format = fmt
        c.fill = RESULT_FILL
        c.font = BOLD
        r += 1
    if price:
        ws[f"A{r}"] = f"Market price ({cur})"
        ws[f"B{r}"] = price
        ws[f"B{r}"].number_format = NUM
        r += 1
        ws[f"A{r}"] = "Upside to base"
        ws[f"B{r}"] = ps / price - 1
        ws[f"B{r}"].number_format = PCT
        ws[f"B{r}"].font = BOLD


# --------------------------------------------------------------------------
# Sheet 2 - Equity DCF (live formulas)
# --------------------------------------------------------------------------

def sheet_equity_dcf(ws, drivers, fc):
    cur = drivers.get("currency", "IDR")
    ws.column_dimensions["A"].width = 28
    for col in "BCDE":
        ws.column_dimensions[col].width = 14

    ws["A1"] = "Equity DCF (live) - edit yellow cells, value recalcs"
    ws["A1"].font = Font(bold=True, size=13)

    coe = drivers["cost_of_equity"]
    cv = drivers["continuing_value"]

    r = 3
    _section(ws, r, "Discount rate", 2); r += 1
    ws[f"A{r}"] = "Risk-free rate"
    ws[f"B{r}"] = coe["risk_free"]; ws[f"B{r}"].fill = INPUT_FILL; ws[f"B{r}"].number_format = PCT
    rf_row = r; r += 1
    ws[f"A{r}"] = "Equity risk premium"
    ws[f"B{r}"] = coe["erp"]; ws[f"B{r}"].fill = INPUT_FILL; ws[f"B{r}"].number_format = PCT
    erp_row = r; r += 1
    ws[f"A{r}"] = "Beta"
    ws[f"B{r}"] = coe.get("beta", 1.0); ws[f"B{r}"].fill = INPUT_FILL; ws[f"B{r}"].number_format = "0.00"
    beta_row = r; r += 1
    ws[f"A{r}"] = "Cost of equity (ke)"
    ws[f"B{r}"] = f"=B{rf_row}+B{beta_row}*B{erp_row}"
    ws[f"B{r}"].number_format = PCT; ws[f"B{r}"].font = BOLD; ws[f"B{r}"].fill = RESULT_FILL
    ke_cell = f"B{r}"; r += 1

    r += 1
    _section(ws, r, "Explicit equity cash flow", 5); r += 1
    _hdr(ws, r, ["Year", "Equity CF", "Period", "Disc factor", "PV"]); r += 1
    first = r
    for i, row in enumerate(fc["rows"], start=1):
        ws.cell(row=r, column=1, value=row["year"])
        c = ws.cell(row=r, column=2, value=round(row["equity_cash_flow"]))
        c.number_format = NUM; c.fill = INPUT_FILL
        ws.cell(row=r, column=3, value=i)
        ws.cell(row=r, column=4, value=f"=1/(1+{ke_cell})^(C{r}-0.5)").number_format = "0.0000"
        ws.cell(row=r, column=5, value=f"=B{r}*D{r}").number_format = NUM
        r += 1
    last = r - 1
    n = len(fc["rows"])
    ws[f"A{r}"] = "PV explicit"; ws[f"A{r}"].font = BOLD
    ws[f"E{r}"] = f"=SUM(E{first}:E{last})"; ws[f"E{r}"].number_format = NUM
    ws[f"E{r}"].font = BOLD; ws[f"E{r}"].fill = RESULT_FILL
    pv_expl = f"E{r}"; r += 2

    _section(ws, r, "Continuing value", 2); r += 1
    ws[f"A{r}"] = "NI next (CV year)"
    ws[f"B{r}"] = round(fc["ni_next"]); ws[f"B{r}"].fill = INPUT_FILL; ws[f"B{r}"].number_format = NUM
    ni_next_row = r; r += 1
    ws[f"A{r}"] = "Terminal growth (g)"
    ws[f"B{r}"] = cv["growth"]; ws[f"B{r}"].fill = INPUT_FILL; ws[f"B{r}"].number_format = PCT
    g_row = r; r += 1
    ws[f"A{r}"] = "RONE"
    ws[f"B{r}"] = cv["rone"]; ws[f"B{r}"].fill = INPUT_FILL; ws[f"B{r}"].number_format = PCT
    rone_row = r; r += 1
    ws[f"A{r}"] = "CV undiscounted"
    ws[f"B{r}"] = f"=B{ni_next_row}*(1-B{g_row}/B{rone_row})/({ke_cell}-B{g_row})"
    ws[f"B{r}"].number_format = NUM
    cv_undisc = f"B{r}"; r += 1
    ws[f"A{r}"] = "Disc factor (year N)"
    ws[f"B{r}"] = f"=1/(1+{ke_cell})^({n}-0.5)"; ws[f"B{r}"].number_format = "0.0000"
    cv_df = f"B{r}"; r += 1
    ws[f"A{r}"] = "PV continuing value"; ws[f"A{r}"].font = BOLD
    ws[f"B{r}"] = f"={cv_undisc}*{cv_df}"; ws[f"B{r}"].number_format = NUM
    ws[f"B{r}"].font = BOLD; ws[f"B{r}"].fill = RESULT_FILL
    pv_cv = f"B{r}"; r += 2

    _section(ws, r, "Value", 2); r += 1
    ws[f"A{r}"] = "Equity value"; ws[f"A{r}"].font = BOLD
    ws[f"B{r}"] = f"={pv_expl}+{pv_cv}"; ws[f"B{r}"].number_format = NUM
    ws[f"B{r}"].font = BOLD; ws[f"B{r}"].fill = RESULT_FILL
    eq_row = r; r += 1
    ws[f"A{r}"] = "Shares outstanding"
    ws[f"B{r}"] = drivers.get("shares_outstanding", 0); ws[f"B{r}"].fill = INPUT_FILL
    ws[f"B{r}"].number_format = "#,##0.00"
    sh_row = r; r += 1
    ws[f"A{r}"] = f"Value per share ({cur})"; ws[f"A{r}"].font = BOLD
    ws[f"B{r}"] = f"=B{eq_row}/B{sh_row}"; ws[f"B{r}"].number_format = NUM
    ws[f"B{r}"].font = BOLD; ws[f"B{r}"].fill = RESULT_FILL


# --------------------------------------------------------------------------
# Sheet 3 - Tornado
# --------------------------------------------------------------------------

def sheet_tornado(ws, torn, price):
    ws.column_dimensions["A"].width = 26
    for col in "BCDE":
        ws.column_dimensions[col].width = 14
    ws["A1"] = "Tornado - value per share at each driver's low / high"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Base value {torn['base_value']:,.0f}  |  ranked by swing (most important first)"
    ws["A2"].font = ITAL

    top = 4
    _hdr(ws, top, ["Driver", "Low ->", "High ->", "Swing"])
    r = top + 1
    first = r
    for row in torn["rows"]:
        ws.cell(row=r, column=1, value=row["label"])
        ws.cell(row=r, column=2,
                value=round(row["low_val"]) if row["low_val"] is not None else None).number_format = NUM
        ws.cell(row=r, column=3,
                value=round(row["high_val"]) if row["high_val"] is not None else None).number_format = NUM
        ws.cell(row=r, column=4, value=round(row["swing"])).number_format = NUM
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BOX
        r += 1
    last = r - 1

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Swing in value per share"
    chart.height = 8
    chart.width = 18
    data = Reference(ws, min_col=4, min_row=top, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, f"F{top}")


# --------------------------------------------------------------------------
# Sheet 4 - Reverse DCF
# --------------------------------------------------------------------------

def sheet_reverse(ws, rev):
    ws.column_dimensions["A"].width = 30
    for col in "BCD":
        ws.column_dimensions[col].width = 14
    ws["A1"] = f"Reverse DCF - driver value implied by price {rev['target_price']:,.0f}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "What the market is pricing in, holding all other drivers at base."
    ws["A2"].font = ITAL

    rate_keys = {"cost_of_equity", "terminal_growth", "nim", "cost_of_credit",
                 "rone", "explicit_growth_peak"}
    top = 4
    _hdr(ws, top, ["Driver", "Base", "Implied", "Implied ROE"])
    r = top + 1
    for row in rev["rows"]:
        ws.cell(row=r, column=1, value=row["label"])
        is_rate = row["key"] in rate_keys
        b = ws.cell(row=r, column=2, value=row["base_in"])
        b.number_format = PCT if is_rate else "0.00"
        imp = row["implied"]
        ic = ws.cell(row=r, column=3, value=imp if imp is not None else "out of range")
        if imp is not None:
            ic.number_format = PCT if is_rate else "0.00"
        if row.get("implied_roe") is not None:
            rc = ws.cell(row=r, column=4, value=row["implied_roe"])
            rc.number_format = PCT
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BOX
        r += 1


# --------------------------------------------------------------------------
# Sheet 5 - Monte Carlo
# --------------------------------------------------------------------------

def sheet_montecarlo(ws, mc):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    for col in "DE":
        ws.column_dimensions[col].width = 13
    ws["A1"] = f"Monte Carlo - {mc['n']:,} simulations"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Cycle factor co-moves rate, NIM, credit cost and growth, plus driver noise."
    ws["A2"].font = ITAL

    rows = [
        ("Mean", mc["mean"]), ("Std dev", mc["stdev"]),
        ("Min", mc["min"]), ("P10", mc["p10"]), ("P25", mc["p25"]),
        ("Median (P50)", mc["p50"]), ("P75", mc["p75"]), ("P90", mc["p90"]),
        ("Max", mc["max"]),
    ]
    r = 4
    _hdr(ws, r, ["Statistic", "Value / share"]); r += 1
    for name, val in rows:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=round(val) if val is not None else None).number_format = NUM
        r += 1
    if mc["prob_undervalued"] is not None:
        ws.cell(row=r, column=1, value=f"P(value >= price {mc['price']:,.0f})").font = BOLD
        c = ws.cell(row=r, column=2, value=mc["prob_undervalued"])
        c.number_format = "0%"; c.font = BOLD; c.fill = RESULT_FILL

    # histogram data (cols D/E), then chart
    htop = 4
    ws.cell(row=htop, column=4, value="Value bin").fill = HEADER_FILL
    ws.cell(row=htop, column=4).font = WHITE
    ws.cell(row=htop, column=5, value="Count").fill = HEADER_FILL
    ws.cell(row=htop, column=5).font = WHITE
    hr = htop + 1
    first = hr
    for b, cnt in zip(mc["hist"]["bins"], mc["hist"]["counts"]):
        ws.cell(row=hr, column=4, value=round(b)).number_format = NUM
        ws.cell(row=hr, column=5, value=cnt)
        hr += 1
    last = hr - 1
    chart = BarChart()
    chart.type = "col"
    chart.title = "Distribution of value per share"
    chart.height = 9
    chart.width = 20
    chart.gapWidth = 20
    data = Reference(ws, min_col=5, min_row=htop, max_row=last)
    cats = Reference(ws, min_col=4, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "G4")


# --------------------------------------------------------------------------

def build(drivers, stress_spec, n, seed):
    fc = driver_bank.build_forecast(drivers)
    result = driver_bank.value_drivers(drivers)
    res = stress.run_all(drivers, stress_spec, n, seed)

    wb = Workbook()
    sheet_drivers(wb.active, drivers, fc, result)
    wb.active.title = "Drivers"
    sheet_equity_dcf(wb.create_sheet("Equity DCF"), drivers, fc)
    sheet_tornado(wb.create_sheet("Tornado"), res["tornado"], drivers.get("price"))
    if res.get("reverse"):
        sheet_reverse(wb.create_sheet("Reverse DCF"), res["reverse"])
    sheet_montecarlo(wb.create_sheet("Monte Carlo"), res["montecarlo"])
    return wb


def main():
    p = argparse.ArgumentParser(description="Build the BBCA stress-test workbook")
    p.add_argument("drivers")
    p.add_argument("--stress", required=True)
    p.add_argument("-o", "--out", default=None)
    p.add_argument("--n", type=int, default=0)
    p.add_argument("--seed", type=int, default=None)
    args = p.parse_args()

    with open(args.drivers) as f:
        drivers = json.load(f)
    with open(args.stress) as f:
        stress_spec = json.load(f)

    wb = build(drivers, stress_spec, args.n, args.seed)
    out = args.out or f"{drivers.get('ticker', 'stress')}_stress.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
