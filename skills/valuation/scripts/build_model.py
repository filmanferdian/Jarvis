"""Build an Excel DCF model from a model.json.

The workbook carries LIVE formulas for the core DCF (WACC, discounting,
continuing value, equity bridge) so the analyst can change an assumption in
Excel and watch value recalc. A second sheet holds a static WACC x growth
sensitivity grid (computed in Python via sensitivity.py).

Requires openpyxl — run with the skill's venv:
    .venv/bin/python scripts/build_model.py <model.json> [-o out.xlsx]
"""

from __future__ import annotations

import argparse
import json

import engine
import sensitivity

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="D6DCE5")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow = analyst input
RESULT_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _label(ws, cell, text, bold=False):
    ws[cell] = text
    if bold:
        ws[cell].font = BOLD


def _input(ws, cell, value, fmt=None):
    ws[cell] = value
    ws[cell].fill = INPUT_FILL
    ws[cell].border = BOX
    if fmt:
        ws[cell].number_format = fmt


def _formula(ws, cell, formula, fmt=None, result=False):
    ws[cell] = formula
    ws[cell].border = BOX
    if result:
        ws[cell].fill = RESULT_FILL
        ws[cell].font = BOLD
    if fmt:
        ws[cell].number_format = fmt


def _section(ws, row, text, span=2):
    c = ws.cell(row=row, column=1, value=text)
    c.font = BOLD
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL


PCT = "0.00%"
NUM = "#,##0.0"
MONEY = "#,##0.00"


def build(model: engine.Model) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16

    # Title
    ws["A1"] = f"{model.company} ({model.ticker}) — Enterprise DCF"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"{model.units} {model.currency} | as of {model.valuation_date}"
    ws["A2"].font = Font(italic=True, color="808080")

    w = model.wacc
    cv = model.continuing_value
    b = model.bridge

    # --- WACC assumptions ---
    r = 4
    _section(ws, r, "Cost of capital")
    rows = {}
    r += 1
    if w.wacc_override is not None:
        _label(ws, f"A{r}", "WACC (override)")
        _input(ws, f"B{r}", w.wacc_override, PCT)
        rows["wacc"] = r
        r += 1
    else:
        _label(ws, f"A{r}", "Risk-free rate")
        _input(ws, f"B{r}", w.risk_free or 0, PCT); rows["rf"] = r; r += 1
        _label(ws, f"A{r}", "Equity risk premium")
        _input(ws, f"B{r}", w.erp or 0, PCT); rows["erp"] = r; r += 1
        _label(ws, f"A{r}", "Beta (levered)")
        _input(ws, f"B{r}", w.beta or 0, "0.00"); rows["beta"] = r; r += 1
        _label(ws, f"A{r}", "Cost of equity")
        if w.cost_of_equity is not None:
            _input(ws, f"B{r}", w.cost_of_equity, PCT)
        else:
            _formula(ws, f"B{r}", f"=B{rows['rf']}+B{rows['beta']}*B{rows['erp']}", PCT)
        rows["ke"] = r; r += 1
        _label(ws, f"A{r}", "Pre-tax cost of debt")
        _input(ws, f"B{r}", w.pretax_cost_of_debt, PCT); rows["kd"] = r; r += 1
        _label(ws, f"A{r}", "Tax rate")
        _input(ws, f"B{r}", w.tax_rate, PCT); rows["tax"] = r; r += 1
        _label(ws, f"A{r}", "Equity weight (E/V)")
        _input(ws, f"B{r}", w.equity_weight, PCT); rows["ew"] = r; r += 1
        _label(ws, f"A{r}", "Debt weight (D/V)")
        _input(ws, f"B{r}", w.debt_weight, PCT); rows["dw"] = r; r += 1
        _label(ws, f"A{r}", "WACC", bold=True)
        _formula(
            ws, f"B{r}",
            f"=B{rows['ew']}*B{rows['ke']}+B{rows['dw']}*B{rows['kd']}*(1-B{rows['tax']})",
            PCT, result=True,
        )
        rows["wacc"] = r; r += 1
    wacc_cell = f"B{rows['wacc']}"

    # --- Explicit FCF ---
    r += 1
    _section(ws, r, "Explicit free cash flow")
    r += 1
    hdr = r
    for col, txt in enumerate(["Year", "FCF", "Period", "Discount factor", "PV"], start=1):
        c = ws.cell(row=hdr, column=col, value=txt)
        c.fill = HEADER_FILL; c.font = WHITE; c.alignment = Alignment(horizontal="center")
    r += 1
    midyear = w.midyear
    first_fcf_row = r
    for i, row in enumerate(model.forecast, start=1):
        ws.cell(row=r, column=1, value=row.get("year"))
        _input(ws, f"B{r}", float(row["fcf"]), NUM)
        ws.cell(row=r, column=3, value=i)
        exp = f"(C{r}-0.5)" if midyear else f"C{r}"
        _formula(ws, f"D{r}", f"=1/(1+{wacc_cell})^{exp}", "0.0000")
        _formula(ws, f"E{r}", f"=B{r}*D{r}", NUM)
        r += 1
    last_fcf_row = r - 1
    n = len(model.forecast)
    _label(ws, f"A{r}", "PV of explicit FCF", bold=True)
    _formula(ws, f"E{r}", f"=SUM(E{first_fcf_row}:E{last_fcf_row})", NUM, result=True)
    pv_explicit_cell = f"E{r}"
    r += 2

    # --- Continuing value ---
    _section(ws, r, "Continuing value")
    r += 1
    _label(ws, f"A{r}", "Method")
    ws[f"B{r}"] = cv.method
    r += 1
    _label(ws, f"A{r}", "Continuing growth (g)")
    _input(ws, f"B{r}", cv.growth, PCT); g_row = r; r += 1
    if cv.method == "key_value_driver":
        _label(ws, f"A{r}", "RONIC")
        _input(ws, f"B{r}", cv.ronic or 0, PCT); ronic_row = r; r += 1
        _label(ws, f"A{r}", "NOPLAT (year N+1)")
        _input(ws, f"B{r}", cv.noplat_next or 0, NUM); noplat_row = r; r += 1
        _label(ws, f"A{r}", "Continuing value (undiscounted)")
        _formula(
            ws, f"B{r}",
            f"=B{noplat_row}*(1-B{g_row}/B{ronic_row})/({wacc_cell}-B{g_row})",
            NUM,
        )
    else:  # perpetuity
        _label(ws, f"A{r}", "FCF (year N+1)")
        _input(ws, f"B{r}", cv.fcf_next or 0, NUM); fcf_next_row = r; r += 1
        _label(ws, f"A{r}", "Continuing value (undiscounted)")
        _formula(ws, f"B{r}", f"=B{fcf_next_row}/({wacc_cell}-B{g_row})", NUM)
    cv_undisc_row = r; r += 1
    _label(ws, f"A{r}", "Discount factor (year N)")
    cv_exp = f"({n}-0.5)" if midyear else f"{n}"
    _formula(ws, f"B{r}", f"=1/(1+{wacc_cell})^{cv_exp}", "0.0000")
    cv_df_row = r; r += 1
    _label(ws, f"A{r}", "PV of continuing value", bold=True)
    _formula(ws, f"B{r}", f"=B{cv_undisc_row}*B{cv_df_row}", NUM, result=True)
    pv_cv_cell = f"B{r}"
    r += 2

    # --- Equity bridge ---
    _section(ws, r, "Equity bridge")
    r += 1
    _label(ws, f"A{r}", "Value of operations", bold=True)
    _formula(ws, f"B{r}", f"={pv_explicit_cell}+{pv_cv_cell}", NUM, result=True)
    op_row = r; r += 1
    _label(ws, f"A{r}", "+ Nonoperating assets")
    _input(ws, f"B{r}", b.nonoperating_assets, NUM); nonop_row = r; r += 1
    _label(ws, f"A{r}", "= Enterprise value", bold=True)
    _formula(ws, f"B{r}", f"=B{op_row}+B{nonop_row}", NUM, result=True)
    ev_row = r; r += 1
    _label(ws, f"A{r}", "- Debt & equivalents")
    _input(ws, f"B{r}", b.debt_and_equivalents, NUM); debt_row = r; r += 1
    _label(ws, f"A{r}", "- Other claims")
    _input(ws, f"B{r}", b.other_claims, NUM); claims_row = r; r += 1
    _label(ws, f"A{r}", "= Equity value", bold=True)
    _formula(ws, f"B{r}", f"=B{ev_row}-B{debt_row}-B{claims_row}", NUM, result=True)
    eq_row = r; r += 1
    _label(ws, f"A{r}", "Shares outstanding")
    _input(ws, f"B{r}", b.shares_outstanding, NUM); sh_row = r; r += 1
    _label(ws, f"A{r}", "Value per share", bold=True)
    _formula(ws, f"B{r}", f"=IF(B{sh_row}=0,\"\",B{eq_row}/B{sh_row})", MONEY, result=True)

    _build_sensitivity_sheet(wb, model)
    return wb


def _build_sensitivity_sheet(wb: Workbook, model: engine.Model) -> None:
    gr = sensitivity.grid(model, metric="per_share", wacc_step=0.005,
                          growth_step=0.005, n=2)
    ws = wb.create_sheet("Sensitivity")
    ws.column_dimensions["A"].width = 14
    ws["A1"] = f"{model.company} — value per share: WACC (rows) x growth (cols)"
    ws["A1"].font = BOLD
    ws["A2"] = "Static snapshot; edit inputs on Valuation sheet and re-run build_model.py to refresh."
    ws["A2"].font = Font(italic=True, color="808080")

    top = 4
    ws.cell(row=top, column=1, value="WACC \\ g").font = BOLD
    for j, g in enumerate(gr["growth_axis"], start=2):
        c = ws.cell(row=top, column=j, value=g)
        c.number_format = PCT; c.font = WHITE; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for i, (wv, vrow) in enumerate(zip(gr["wacc_axis"], gr["values"]), start=top + 1):
        rc = ws.cell(row=i, column=1, value=wv)
        rc.number_format = PCT; rc.font = WHITE; rc.fill = HEADER_FILL
        for j, val in enumerate(vrow, start=2):
            c = ws.cell(row=i, column=j, value=val)
            c.number_format = MONEY
            c.border = BOX
            is_base = (abs(wv - gr["base_wacc"]) < 1e-9 and
                       abs(gr["growth_axis"][j - 2] - gr["base_growth"]) < 1e-9)
            if is_base:
                c.fill = RESULT_FILL; c.font = BOLD


def main() -> None:
    p = argparse.ArgumentParser(description="Build an Excel DCF model from model.json")
    p.add_argument("model")
    p.add_argument("-o", "--out", default=None, help="output .xlsx path")
    args = p.parse_args()

    with open(args.model) as f:
        data = json.load(f)
    model = engine.Model.from_dict(data)
    wb = build(model)

    out = args.out or f"{model.ticker or 'valuation'}_dcf.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
